import os
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import re
import time
from data import areas_info, work_dicts
import db
import uuid
import xlrd
import datetime
import xlrd.biffh
import string

# punct = знаки пунктуации (чтобы исключить возможность заголовков с пунктуацией)
punct = string.punctuation


def get_clean_num(num, to_check=None):
    if to_check is False:
        pu_num = '0'
        try:
            pu_num = re.sub(r'\D', '', num)
        except Exception as e:
            pass
            # print(repr(e))
        return pu_num


def get_area(data):
    # areas_info[key] = [res_area_name, res_id, id_area]
    res_id, area_id, area_name = '', '', ''
    for area_entry in data:
        if area_entry:
            if 'Московский' in area_entry:
                area_entry = 'Московский РЭС'
            if 'Ленинский' in area_entry:
                area_entry = 'Ленинский РЭС'
            if 'Серпуховской' in area_entry:
                area_entry = 'Серпуховский РЭС'
            area_entry = area_entry.strip().lower()
            try:
                area_name = work_dicts['areas'][area_entry][0]
                res_id = work_dicts['areas'][area_entry][1]
                area_id = work_dicts['areas'][area_entry][2]
            except Exception as e:
                pass
                # print('get_area', repr(e), type(e))
    return res_id, area_id, area_name


# ф-ция для формирования dict по данным БД,
# для дальнейшего сравнения с данным из Excel в модуле data
def area_to_res():
    work_dicts['areas'] = {}
    # пример result  = ['Балашихинский городской округ', id],...
    result = db.qdb(query="""SELECT area_name, id
                                FROM ENERSTROYMAIN_area 
                            where zone_id in('0af04d9f-bc63-90a9-afcc-8cc54c8d8c3d',
                                'a07b5493-df1f-d125-ed00-4e5cdc39b83e')
                            and DELETE_TS is null;""", debug_status=debug_status)
    area_name = None
    if result:
        for row in result:
            area_name_db = row[0]  # обычная запись = 'Балашиха городской округ'
            area_id = row[1]  # id='f06f7420-0ea2-4a96-a970-3a554b0f72ad'
            # areas_info:
            # {'Балашиха городской округ':
            #                       ['Балашихинский РЭС',
            #                        'f06f7420-0ea2-4a96-a970-3a554b0f72ad',
            #                        'e98626e1-8b2d-f917-455e-c127576c1cf3'],...}
            for area_name_table in areas_info:
                if area_name_table == area_name_db and area_id == areas_info[area_name_table][2]:
                    area_name = area_name_db.lower().strip()
                    # area_res_name = 'балашихинский рэс'
                    area_res_name = areas_info[area_name_table][0].lower().strip()
                    res_id = areas_info[area_name_table][1]
                    # примеры данных выше:
                    # areas_info[area_name_db][0] = 'балашихинский рэс' - area_name
                    # areas_info[area_name_db][2] = 'f06f7420-0ea2-4a96-a970-3a554b0f72ad' = res_id
                    # areas_info[area_name_db][1] = 'f06f7420-0ea2-4a96-a970-3a554b0f72ad' = area_id
                    work_dicts['areas'][area_res_name] = [area_name, res_id, area_id]
        print(work_dicts['areas'])


# Ф-ция для формирования шаблона сравнения с данными из Excel
def gen_fiz_keys():
    work_dicts['fiz_keys'] = {}
    # Было:
    # row[0] = accounting_num, row[1] = pu_num, row[2] = id
    # row[3] = name, row[4] = date_kp1
    # row[5] = date_kp2, row[6] = poselenie, row[7] = pu_type

    # '''SELECT accounting_num, pu_num, id,
    #                         name, date_kp1, date_kp2, poselenie, pu_type
    #                    FROM public.enerstroymain_subscriber
    #                    where delete_ts is null;'''

    # Стало:
    # row[0] = subscriber_accounting_num, row[1] = subscriber_pu_num, row[2] = subscriber_id,
    # row[3] = subscriber_address_table_id, row[4] = subscriber_name, row[5] = subscriber_date_kp1
    # row[6] = subscriber_date_kp2, row[7] = subscriber_poselenie, row[8] = subcriber_pu_type
    # row[9] = address_table_LOCALITY_NAME_MOB, row[10] = address_table_STREET_NAME_MOB,
    # row[11] = address_table_HOUSE, row[12] = address_table_APARTMENT

    result = db.qdb(r"""SELECT sub.accounting_num, sub.pu_num, sub.id, sub.ADDRESS_TABLE_ID,
                            sub.name, sub.date_kp1, sub.date_kp2, sub.poselenie, sub.pu_type,
                            addr.LOCALITY_NAME_MOB, addr.STREET_NAME_MOB, addr.HOUSE, addr.APARTMENT 
                       FROM public.enerstroymain_subscriber sub
                       left join ENERSTROYMAIN_ADDRESS_TABLE addr on addr.id=sub.ADDRESS_TABLE_ID
                       where sub.delete_ts is null;""")
    if result:
        for row in result:
            # если формат данных не время выставляем Null
            datapk1, datapk2 = 'Null', 'Null'
            if type(row[5]) == datetime.date:
                datapk1 = row[5].strftime("%Y-%m-%d")
            elif row[5] is None:
                pass

            if type(row[6]) == datetime.date:
                datapk2 = row[6].strftime("%Y-%m-%d")
            elif row[6] is None:
                pass
            # по новой логике ls
            # должен быть integer без лишних символов
            # даже если ls не указан в БД
            # мы должны составить dict c 0 в формате int
            ls_int = 0
            ls_raw = row[0]
            try:
                ls = re.sub(r'\D', '', ls_raw)
                ls_int = int(ls)
            except (ValueError, TypeError):
                # проверочный принт
                # print(ls_raw, ls_int)
                pass

            # по новой логике pu_num
            # должен быть integer без лишних символов,
            # даже если pu_num не указан в БД
            # мы должны составить dict c 0 в формате int
            pu_num_int = 0
            pu_num_raw = row[1]
            try:
                pu_num = re.sub(r'\D', '', pu_num_raw)
                pu_num_int = int(pu_num)
            except (ValueError, TypeError):
                # проверочный принт
                # print(pu_num_raw, pu_num_int)
                pass

            # проверка есть ли в fiz_keys accounting_num
            if ls_int not in work_dicts['fiz_keys']:
                # если и нет создаем пукстой словарь по этому ключю
                work_dicts['fiz_keys'][ls_int] = {}
                # инициализируем ключи и записываем данные:
                # 'repeat' =  повторы в базе данных
                # 'is_inserted' = колличество инсертов
                # 'updated' = колличество апдейтов
                # 'data_subscriber' : [ id, ADDRESS_TABLE_ID,
                #                             name, date_kp1, date_kp2, poselenie, pu_type,]

                # 'data_subscriber':[addr.LOCALITY_NAME_MOB, addr.STREET_NAME_MOB,
                #                       addr.HOUSE, addr.APARTMENT]
                work_dicts['fiz_keys'][ls_int][pu_num_int] = {'data_subscriber': [row[2], row[3], datapk1, datapk2,
                                                                                  row[6], row[7], row[8]],
                                                              'data_address': [row[9], row[10], row[11], row[12]],
                                                              'is_address_table_id_updated': False,
                                                              'repeat': 0,
                                                              'is_inserted': 0,
                                                              'updated': 0}
            # если есть такой ls_int
            else:
                # проверка есть ли pu_num_int у такого ls_int
                if pu_num_int not in work_dicts['fiz_keys'][ls_int]:
                    work_dicts['fiz_keys'][ls_int][pu_num_int] = {
                        'data_subscriber': [row[2], row[3], datapk1, datapk2, row[6], row[7], row[8]],
                        'data_address': [row[9], row[10], row[11], row[12]],
                        'is_address_table_id_updated': False,
                        'repeat': 0, 'is_inserted': 0, 'updated': 0}
                else:
                    work_dicts['fiz_keys'][ls_int][pu_num_int]['repeat'] += 1

    print("fiz_keys's length", f'{len(work_dicts["fiz_keys"])}')


# Проверка значения в excel(является ли значение заголовокм пример: NPEC)
def check_header(raw_value):
    status_header = False
    if len(raw_value) > 0:
        for letter in raw_value:
            if letter.isupper():
                status_header = True
            elif letter in punct:
                pass
            else:
                status_header = False
        return status_header
    else:
        return None


# ф-ция для считывания данных из excel сравнения значений с шаблоном из БД
# принятие решения об insert'е или update'e, вычисление дубликатов в БД,
# сверка данных на актуальность
def read_xlsx(file):
    wb = load_workbook(filename=f'{work_dicts["work_dir"]}/{file}')
    ws = wb.active
    created_by = 'patrik'
    updated_by = 'patrik'
    addr_condition = "Null"
    addrobj_id = "Null"
    # doubles = повторяющиеся в accounting_num-pu_num в БД
    doubles = 0
    # count = строка по которой происходит итерация.
    # т.к. нумерация в excel начинается в excel файле начинается с 1
    # count изанчально = 1
    count = 1
    # error = строка, где информация некорректна
    errors = 0
    # suc = строка, по значениям которой
    # произведен успешно insert
    suc = 0
    # in_base = значения строки,
    # по которой производится итерация,
    # аналогична значениям в БД
    in_base = 0
    # updated = изменения внесены в БД
    updated = 0
    # колличество проанализированных строк
    row_counter = 0
    load_type = 'old'
    # Если верхние строчки смерджины(объединены в одну)
    if type(ws.cell(row=count, column=count + 1)).__name__ == 'MergedCell':
        count += 1
    # check if row value is header
    # добавлена проверка названия столбца и его значения
    is_header = check_header(ws.cell(row=count, column=40).value)
    if is_header is True:
        count += 1
    for row in ws.iter_rows(min_row=count):
        try:
            # ls = NPOTERI_PR- в excel(пример),  Accounting_Num - в БД
            ls = str(row[40].value)
            # сканируем занчение в 40-овом стобце для  поддержки старой логики
            if not ls or len(ls) < 12 or not re.search('\d{5}-\d{3}-\d{2}', ls):
                # новая логика для = проверка поля IKTS(10-го стобца)
                # IKTS - в Excel, Accounting_Num - в БД
                ls = str(row[10].value)  # IKTS
                if not ls or len(ls) < 12 or not re.search('\d{5}-\d{3}-\d{2}', ls):
                    if ls is not None and row[40].value is not None:
                        ws.cell(column=42, row=count, value="Не корректный ЛС")
                        errors += 1
                        count += 1
                        row_counter += 1
                        continue
            else:
                if row[40].value is not None and row[10].value is not None:
                    ws.cell(column=42, row=count, value="Не удалось распарсить ЛС")
                    errors += 1
                    count += 1
                    row_counter += 1
                    continue

            # по новой логике если не получилось
            # переконвертировать ls в integer
            # то должны возвратить 0
            ls_int = 0
            try:
                ls_str = re.sub(r'\D', '', ls)
                ls_int = int(ls_str)
            except Exception as e:
                pass
                # print(repr(e))

            load_type = 'new'
            # pu_num = PU_NUM(в БД)
            # pu_num = CHNUM(в эксель таблицах)
            # pu_num_int - по новой логике мы записываем именно значение в формате string
            # это может быть 0, если chnum не указан
            pu_num_int = int(get_clean_num(row[16].value, to_check=False))
            if pu_num_int is None or row[16].value is None:
                if row[40].value is None or row[10].value is None:
                    break
                elif pu_num_int is None:
                    ws.cell(column=42, row=count, value="Не корректный номер ПУ")
                    errors += 1
                    count += 1
                    row_counter += 1
                    continue

            # area_id = NMPEC column, area_name= NMELSETI_M column в Excel
            # row[2].value = NMELSETI_M column, row[3].value = NNELSETI_M column
            res_id, area_id, area_name = get_area([row[2].value, row[3].value])
            if not area_id:
                ws.cell(column=42, row=count, value="Не удалось опеределить РЭС")
                count += 1
                errors += 1
                row_counter += 1
                continue

            # формирование данных для БД
            id = uuid.uuid4()  # id в БД
            address_table_id = uuid.uuid4()  # генерация id-шника для address_table_id
            DATAPK1, DATAPK2, PU_TYPE = 'Null', 'Null', 'Null'  # если не получится привести занчения
            # к нужному формату выставляем в Null
            name = row[11].value  # FIO - в Excel, в name - В БД
            POSELENIE = row[12].value  # POSELENIE в Excel и в БД
            STREET = row[13].value  # STREET в Excel и в БД
            HOUSE = row[14].value  # HOUSE в Excel и в БД
            FLATNUM = row[15].value  # FLATNUM в Excel и в БД
            zone_name = row[1].value  # NMPEC в Excel и в БД
            # попытка привести данные к нужному формату, если не получается то, Null
            if load_type == 'old':
                try:
                    DATAPK1 = row[26].value.strftime("%Y-%m-%d")
                except Exception as e:
                    pass
                try:
                    DATAPK2 = row[27].value.strftime("%Y-%m-%d")
                except Exception as e:
                    pass

            elif load_type == 'new':
                try:
                    DATAPK1 = row[34].value.strftime("%Y-%m-%d")
                except Exception as e:
                    pass
                try:
                    DATAPK2 = row[35].value.strftime("%Y-%m-%d")
                except Exception as e:
                    pass
            # raw в конце означает сырой формат данных, нужно привести к нужному
            PU_UST_raw = row[20].value
            PU_UST = None
            # проверка на отсутсвие значения PU_UST_raw
            if PU_UST_raw is not None:
                PU_UST_lst = str(PU_UST_raw).split('.')
                for idx, entry_pu_ust in enumerate(PU_UST_lst):
                    if entry_pu_ust == '.':
                        PU_UST_lst[idx] = '-'
                # приводим к формату 2020-07-10, Year-month-day
                PU_UST = '-'.join(PU_UST_lst[::-1])
            else:
                # если значение в столбце не указано, то пишем Null
                PU_UST = 'Null'
            PU_TYPE = row[17].value
            # full address reversed проверчное поле, которое добавляется в БД
            full_address_reversed = f'''кв.{FLATNUM},д.{HOUSE},{STREET},{POSELENIE},{area_name},{zone_name}'''
            print('full address reversed', full_address_reversed, count)
            # провекрка ls в work_dicts['fiz_keys'], проверка pu_num'а, принадлежащего этому ls
            # а так же проверка на дубли(repeat) связки в БД,
            # и проверка: insert'ился ли такая связка ранее в данном файле.
            if ls_int in work_dicts['fiz_keys'] and pu_num_int in work_dicts['fiz_keys'][ls_int] and \
                    work_dicts['fiz_keys'][ls_int][pu_num_int]['repeat'] == 0 and \
                    work_dicts['fiz_keys'][ls_int][pu_num_int]['is_inserted'] == 0:
                # здесь создаем address_table для этого ls и pu_num.

                # как и в случае с update'ом мы должны вставитьь adress_condition=77
                # и addrobj_id = 'Null'

                # сверка с шаблоном есть ли такие даннные в БД или есть ли изменения
                # 'data_subscriber':[ id, ADDRESS_TABLE_ID, name, date_kp1, date_kp2, poselenie, pu_type,]
                if (name != work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'][2]) or (
                        POSELENIE != work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'][5]) or (
                        PU_TYPE != 'Null' and PU_TYPE != work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'][6]) or (
                        DATAPK1 != 'Null' and DATAPK1 != work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'][3]) or (
                        DATAPK2 != 'Null' and DATAPK2 != work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'][4]) or (
                        work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'][1] == 'Null'):

                    sql_insert_addresses = f"""INSERT INTO ENERSTROYMAIN_ADDRESS_TABLE(HOUSE,LOCALITY_NAME_MOB,
                                                                                       CREATE_TS,ID,AREA_ID,UPDATED_BY,
                                                                                       VERSION,CREATED_BY,UPDATE_TS,
                                                                                       STREET_NAME_MOB,APARTMENT,
                                                                                       addrobj_id, address_condition)
                                                                                   VALUES('%s','%s',now(),'%s','%s','%s',
                                                                                   1,'%s', now(),'%s','%s', '%s', '%s');""" \
                    % (HOUSE, POSELENIE, id, area_id, updated_by, created_by, STREET, FLATNUM, addrobj_id, addr_condition)
                    crt_sql_addr_ins = db.clear_sql(sql_insert_addresses)
                    result_upd_sub = db.qdb(crt_sql_addr_ins, type_status=3, debug_status=debug_status)
                    if result_upd_sub == 1:
                        ws.cell(column=46, row=count,
                        value="Для этого IKTS-CHNUM создана address_table, addrobj_id=Null, address_condition=77 ")
                    else:
                        ws.cell(column=46, row=count,
                                value="Для этого IKTS-CHNUM не создана address_table")

                    upd_sql = f"""update ENERSTROYMAIN_SUBSCRIBER set
                                    PU_NUM = '%s',NAME='%s',POSELENIE='%s',STREET='%s',
                                    HOUSE='%s',FLAT_NUM='%s',area_id='%s',
                                    full_address_reversed='%s',DATE_KP1='%s',
                                    DATE_KP2='%s',PU_TYPE='%s',
                                    UPDATED_BY='%s',UPDATE_TS = now(), pu_montage_date='%s', addres_table_id='%s'
                                where id='%s';""" \
                              % (pu_num_int, name, POSELENIE, STREET, HOUSE, FLATNUM, area_id,
                                 full_address_reversed,
                                 DATAPK1, DATAPK2, PU_TYPE, updated_by, PU_UST, address_table_id,
                                 work_dicts['fiz_keys'][ls_int][pu_num_int]['data'][0])

                    upd_sql = db.clear_sql(upd_sql)
                    result_upd_sub = db.qdb(upd_sql, type_status=3, debug_status=debug_status)
                    if result_upd_sub == 1:
                        if work_dicts['fiz_keys'][ls_int][pu_num_int]['updated'] == 0:
                            ws.cell(column=42, row=count,
                                    value="Такой ЛС уже существует, остальная информация обновлена")
                            # выставляем флаг говорящий о том, что addres_table_id у этого ls-pu_num обновлен
                            work_dicts['fiz_keys'][ls_int][pu_num_int]['is_address_table_id_updated'] = True
                            work_dicts['fiz_keys'][ls_int][pu_num_int]['updated'] += 1
                        else:
                            ws.cell(column=42, row=count,
                                    value=f"Такой ЛС уже существует, остальная информация обновлялась {work_dicts['fiz_keys'][ls_int][pu_num_int]['updated']} ")
                            work_dicts['fiz_keys'][ls_int][pu_num_int]['updated'] += 1

                        work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'] = [id, name, DATAPK1,
                                                                                         DATAPK2, POSELENIE,
                                                                                         PU_TYPE]
                        updated += 1
                        ws.cell(column=43, row=count,
                                value='%s' % work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'][0])

                    else:
                        ws.cell(column=42, row=count, value="Не записано")
                        errors += 1
                else:
                    ws.cell(column=42, row=count, value="Запись по ЛС в йоде актуальна")
                    ws.cell(column=43, row=count,
                            value='%s' % work_dicts['fiz_keys'][ls_int][pu_num_int]['data_subscriber'][0])
                    in_base += 1

                count += 1
                row_counter += 1

            elif ls_int in work_dicts['fiz_keys'] and pu_num_int in work_dicts['fiz_keys'][ls_int] and \
                    work_dicts['fiz_keys'][ls_int][pu_num_int]['repeat'] != 0:
                ws.cell(column=42, row=count,
                        value="""В базе данных пристутсвует дубль связки IKTS-CHNUM, обратитесь в тех поддержку""")
                count += 1
                row_counter += 1
                doubles += 1

            elif ls_int in work_dicts['fiz_keys'] and pu_num_int in work_dicts['fiz_keys'][ls_int] and \
                    work_dicts['fiz_keys'][ls_int][pu_num_int]['is_inserted'] != 0:
                ws.cell(column=42, row=count,
                        value="""Запись с таким IKTS-CHNUM уже была добавлены в йоду""")
                count += 1
                row_counter += 1
                doubles += 1

            else:
                sql_insert_addresses = f"""INSERT INTO ENERSTROYMAIN_ADDRESS_TABLE(HOUSE,LOCALITY_NAME_MOB,
                                                    CREATE_TS,ID,AREA_ID,UPDATED_BY,VERSION,CREATED_BY,UPDATE_TS,
                                                    STREET_NAME_MOB,APARTMENT)
                                                VALUES('%s','%s',now(),'%s','%s','%s',
                                                1,'%s', now(),'%s','%s');""" % (
                    HOUSE, POSELENIE, id, area_id, updated_by, created_by, STREET, FLATNUM)

                sql_insert_subscriber = f"""INSERT INTO ENERSTROYMAIN_SUBSCRIBER
                                                (CALC_COEFFICIENT, version,CREATED_BY,
                                                CREATE_TS,UPDATED_BY,UPDATE_TS,ID,
                                                ACCOUNTING_NUM,PU_NUM,NAME,
                                                POSELENIE,STREET,HOUSE,
                                                FLAT_NUM,area_id,DATE_KP1,
                                                DATE_KP2,PU_TYPE,PU_MONTAGE_DATE,
                                                RES_ID,ADDRESS_TABLE_ID)
                                            VALUES (1, 1,'%s',now(),'%s',now(),'%s','%s',
                                            '%s','%s','%s','%s','%s','%s','%s','%s','%s',
                                            '%s','%s','%s','%s');""" \
                                        % (created_by, updated_by, id, ls, pu_num_int, name, POSELENIE, STREET, HOUSE,
                                           FLATNUM, area_id, DATAPK1, DATAPK2, PU_TYPE, PU_UST, res_id,
                                           address_table_id)

                sql_ins_addr = db.clear_sql(sql_insert_addresses)
                result_ins_addr = db.qdb(sql_ins_addr, 3, debug_status)
                if result_ins_addr == 1:
                    ws.cell(column=44, row=count, value="Запись по IKTS-CHNUM добавлена в йоду в addresses table")
                    ws.cell(column=45, row=count, value=f'{id}')
                    # создаем ключ и его значения
                    work_dicts['fiz_keys'][ls_int] = {}
                    work_dicts['fiz_keys'][ls_int][pu_num_int] = {
                        'data_subscriber': [id, name, DATAPK1, DATAPK2, POSELENIE, PU_TYPE],
                        'repeat': 0, 'is_inserted': 0, 'updated': 0}
                    # is_inserted = указатель на то, что данная запись уже была добавлена в йоду
                    work_dicts['fiz_keys'][ls_int][pu_num_int]['is_inserted'] += 1
                    suc += 1
                else:
                    ws.cell(column=43, row=count, value="Не записано")
                    errors += 1

                sql_ins_sub = db.clear_sql(sql_insert_subscriber)
                result_ins_sub = db.qdb(sql_ins_sub, type_status=3, debug_status=debug_status)
                if result_ins_sub == 1:
                    ws.cell(column=42, row=count, value="Запись по IKTS-CHNUM добавлена в йоду в subscriber table")
                    ws.cell(column=43, row=count, value=f'{id}')
                    # при инсерте в addres_table ранее
                    # был создан ключ, но если бы попали в Error
                    # то должны сделать провреку
                    if ls_int not in work_dicts['fiz_keys']:
                        work_dicts['fiz_keys'][ls_int] = {}
                        if pu_num_int not in work_dicts['fiz_keys'][ls_int]:
                            work_dicts['fiz_keys'][ls_int][pu_num_int] = {
                                'data_subscriber': [id, name, DATAPK1, DATAPK2, POSELENIE, PU_TYPE],
                                'repeat': 0, 'is_inserted': 0, 'updated': 0}
                            work_dicts['fiz_keys'][ls_int][pu_num_int]['is_inserted'] += 1
                    suc += 1
                else:
                    ws.cell(column=42, row=count, value="Не записано")
                    errors += 1

                count += 1
                row_counter += 1

        except Exception as e:
            # проверочный принт
            # print(repr(e))
            pass

    wb.save(
        filename=f'{work_dicts["result_dir"]}/errors={errors}_rows={row_counter}_suc={suc}_upd={updated}_in_base={in_base}_doubles={doubles}_{file}')
    try:
        os.rename(f'{work_dicts["work_dir"]}/{file}', f'{work_dicts["finish_dir"]}/{file}')
    except Exception as e:
        # проверовчный принт
        # print(repr(e))
        print(f'move to finish dir not suc {file}')


# ф-ция для конвертации в xlsx файл
def get_xlsx(src_file_path):
    try:
        book_xls = xlrd.open_workbook(f'{work_dicts["work_dir"]}/{src_file_path}')
    except xlrd.biffh.XLRDError:
        return 'file cannot be converted'

    book_xlsx = Workbook()
    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0, len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)
    dst_file_path = f'{work_dicts["work_dir"]}/{src_file_path}x'
    book_xlsx.save(dst_file_path)
    try:
        os.rename(f'{work_dicts["work_dir"]}/{src_file_path}',
                  f'{work_dicts["finish_dir"]}/{src_file_path}')
        return f'{src_file_path}'
    except Exception as e:
        # проверочный принт, не удалось переименовать
        # print(e)
        return None


def check_type_file():
    # проходимся по всем папкам, которые лежат в директории work
    files = os.listdir(work_dicts['work_dir'])
    for file in files:
        # len(work_dicts['fiz_keys']) == length of a list
        if len(work_dicts['fiz_keys']) < 3:
            print(f'недостаточно строк в йоде f{file}')

        print(f'fiz_keys={len(work_dicts["fiz_keys"])}')

        file_name_lst = file.split('.')
        print(file_name_lst)
        # extrude file_type of a file
        file_type = file_name_lst[:-1]

        if file_type == 'xlsb':
            print(f'file {file} not support ')
        # пытаемся переконвертировать в xlsx
        elif 'xlsx' not in file:
            file = get_xlsx(file)
            print(file)

        if file is not None:
            print('inside if file is not None')
            read_xlsx(file)


def start():
    area_to_res()
    if debug_status is False:
        gen_fiz_keys()
    print('start check_type_file')
    check_type_file()
    print('end check_type_file')


if __name__ == '__main__':
    start_t = time.time()
    debug_status = False
    # соединение с БД
    is_connected = db.connect_db()
    if is_connected is True:
        start()
    else:
        print('No connection to DB')
    end_t = time.time()
    print(end_t - start_t)
